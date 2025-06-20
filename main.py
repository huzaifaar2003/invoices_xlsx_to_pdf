import pandas as pd
import fpdf
import glob
import pathlib
import openpyxl
import os
filepaths = glob.glob("invoices_excel/*.xlsx")

print(filepaths)

for filepath in filepaths:

    # sheet_name because an excel file can have multiple sheets

    # extracting invoice number
    filename = pathlib.Path(filepath).stem
    invoice_no = filename.split("-")[0]
    date_list = filename.split("-")[1].split(".")

    # create a pdf corresponding to each excel invoice
    pdf = fpdf.FPDF(orientation="P", unit='mm', format="A4")
    pdf.add_page()
    pdf.set_font(family="Times",style="B", size=20)
    pdf.cell(w=0, h=20, txt=f"Invoice No# {invoice_no}", ln=1)
    pdf.cell(w=0, h=16, txt=f"Date: {date_list[2]}/{date_list[1]}/{date_list[0]}", ln=2)


    df=pd.read_excel(filepath)
    total = df["total_price"].sum()
    headers = [item.replace("_", " ").title() for item in df.columns] # list comprehension
    # add table headers
    pdf.set_font(family="Times", size=10)

    pdf.cell(w=25, h=10, txt=headers[0], border=1)
    pdf.cell(w=45, h=10, txt=headers[1], border=1)
    pdf.cell(w=60, h=10, txt=headers[2], border=1)
    pdf.cell(w=35, h=10, txt=headers[3], border=1)
    pdf.cell(w=20, h=10, txt=headers[4], border=1, ln=1)


    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=10)
        pdf.cell(w=25, h=10, txt=str(row["product_id"]), border=1)
        pdf.cell(w=45, h=10, txt=str(row["product_name"]), border=1)
        pdf.cell(w=60, h=10, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=35, h=10, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=20, h=10, txt=str(row["total_price"]), border=1, ln=1)

    pdf.cell(w=25, h=10, txt="", border=1)
    pdf.cell(w=45, h=10, txt="", border=1)
    pdf.cell(w=60, h=10, txt="", border=1)
    pdf.cell(w=35, h=10, txt="", border=1)
    pdf.cell(w=20, h=10, txt=str(total), border=1, ln=1)
    
    pdf.ln(2)
    pdf.set_font(family="Times", style="ib", size=10)
    pdf.cell(w=20, h=10, txt=f"Total amount due is: -/{total} Dorral")

    pdf.image("r.png", x=10, y= 100, w= 10, type="png")

    # create directory after checking its existence
    if os.path.isdir("PDFs") == False:
        os.mkdir("PDFs")

    pdf.output(f"PDFs/{filename}.pdf")


